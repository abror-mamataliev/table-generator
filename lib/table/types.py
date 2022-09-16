from datetime import date, datetime
from json import load
from openpyxl import load_workbook
from openpyxl.descriptors import Sequence
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any, Union

from lib.sqlite import SQLite
from .exceptions import (
    TableAttributeError,
    TableValueError
)


class Table:

    def __init__(
        self,
        type: str = None,
        table_data: str = None,
        **kwargs
    ) -> None:
        if type is None:
            raise TableAttributeError("Table type is not defined")
        if table_data is None:
            raise TableAttributeError("Table data is required.")
        self._type: str = type
        self._table_data: str = table_data
        for key, value in kwargs.items():
            self.__dict__[f'_{key}'] = value
        self._table_data: dict = self._get_table_data()
    
    def _generate_excel(self, headers: list, data: list) -> str:
        pass

    def _generate_html(self, headers: list, data: list) -> str:
        table = "<table class='table table-bordered text-center'>"
        table += "<thead>"
        for row in headers:
            table += "<tr>"
            for header in row:
                table += f"<th colspan='{header['colspan']}' rowspan='{header['rowspan']}'>{header['name']}</th>"
            table += "</tr>"
        table += "</thead>"
        table += "<tbody>"
        for key in data:
            table += "<tr>"
            for i in range(len(data[key])):
                value = data[key][i]
                if i == 0 and key == 'total':
                    table += f"<td colspan='2'>{value}</td>"
                else:
                    table += f"<td>{round(value, 1) if isinstance(value, float) else value if value is not None else ''}</td>"
            table += "</tr>"
        table += "</tbody>"
        table += "</table>"
        return table

    def _generate_json(self, headers: list, data: list) -> dict:
        new_data: dict = {}
        index: int = 1
        for key in data:
            if key == 'total':
                continue
            new_data[str(index)] = data[key]
            index += 1
        new_data['total'] = data['total']
        return {
            'headers': headers,
            'data': new_data
        }

    def _get_columns(self) -> list:
        return self._table_data['columns']

    def _get_column_keys(self) -> dict:
        return self._table_data['column_keys']
    
    def _get_column_ranges(self) -> dict:
        # merged_cells: list = self._get_merged_cells()
        # column_ranges: dict = {
        #     'min': None,
        #     'max': None
        # }
        # workbook: Workbook = load_workbook(f"utils/excel/{self._type}.xlsx")
        # worksheet: Worksheet = workbook["Жами"]
        # columns = worksheet.columns
        # for cell in merged_cells:
        #     if str(cell).startswith("A1:"):
        #         column_ranges['min'], column_ranges['max'] = [item[0] for item in cell.split(":")]
        # # print(self._table_data['column_ranges'])
        # workbook.close()
        # return column_ranges
        return self._table_data['column_ranges']

    def _get_data(self) -> dict:
        columns: list = self._get_columns()
        column_keys: list = self._get_column_keys()
        data: list = {}
        sql: SQLite = SQLite(f"database/excel/{self._type}.db")
        if '_region' in self.__dict__  and self._region != "all":
            region: str = sql.select(
                table="region",
                fields=["name"],
                where=f"id = {self._region}",
                first=True
            )[0]
            districts: list = sql.select(
                table="data",
                fields=["name", "district", "id"],
                where=f"region = '{region}'",
                group_by="district",
                order_by="id",
                first=False
            )
            for i in range(len(districts)):
                district = districts[i]
                row = []
                for column in columns:
                    type = column['type']
                    if type == "index":
                        row.append(i + 1)
                    elif type == "name":
                        row.append(district[0])
                    else:
                        row.append(0)
                data[district[1]] = row
            data['total'] = []
            for column in columns:
                type = column['type']
                if type == "index":
                    continue
                elif type == "name":
                    data['total'].append("Жами")
                else:
                    data['total'].append(0)
            for i in range(len(districts)):
                district = districts[i]
                result: list = sql.select(
                    table="data",
                    fields=["key", "value"],
                    where=f"region = '{region}' AND district = '{district[1]}' AND name = '{district[0]}' and date = '{self._date}'",
                    first=False
                )
                for item in result:
                    data[district[1]][column_keys[item[0]]] = item[1]
                    if data['total'][column_keys[item[0]] - 1] is None:
                        data['total'][column_keys[item[0]] - 1] = item[1]
                    else:
                        data['total'][column_keys[item[0]] - 1] += item[1]
        else:
            regions: list = sql.select(
                table="region",
                order_by="id",
                first=False
            )
            for i in range(len(regions)):
                region = regions[i]
                row = []
                for column in columns:
                    type = column['type']
                    if type == "index":
                        row.append(i + 1)
                    elif type == "name":
                        row.append(region[1])
                    else:
                        row.append(0)
                data[region[0]] = row
            data['total'] = []
            for column in columns:
                type = column['type']
                if type == "index":
                    continue
                elif type == "name":
                    data['total'].append("Жами")
                else:
                    data['total'].append(0)
            for i in range(len(regions)):
                region = regions[i]
                result: list = sql.select(
                    table="data",
                    fields=["key", "SUM(value)"],
                    where=f"region = '{region[1]}' and date = '{self._date}'",
                    group_by="key",
                    first=False
                )
                for item in result:
                    data[region[0]][column_keys[item[0]]] = item[1]
                    if data['total'][column_keys[item[0]] - 1] is None:
                        data['total'][column_keys[item[0]] - 1] = item[1]
                    else:
                        data['total'][column_keys[item[0]] - 1] += item[1]
        sql.close()
        for key in data:
            for i in range(len(data[key]) - 1, -1, -1):
                if key != 'total':
                    if columns[i]['type'] == "sum":
                        data[key][i] = sum(data[key][j] for j in columns[i]['indexes'])
                        if data['total'][i - 1] is None:
                            data['total'][i - 1] = data[key][i]
                        else:
                            data['total'][i - 1] += data[key][i]
                index: int = i + 1 if key == 'total' else i
                if columns[index]['type'] == "diff":
                    first, second = columns[index]['indexes']
                    data[key][i] = data[key][first - (1 if key == 'total' else 0)] - data[key][second - (1 if key == 'total' else 0)]
        return data

    def _get_headers(self) -> list:
        return self._table_data['headers']

    def _get_header_data(self) -> list:
        workbook: Workbook = load_workbook(f"utils/excel/{self._type}.xlsx")
        worksheet: Worksheet = workbook["Жами"]
        merged_cells = self._get_merged_cells()
        column_ranges = self._get_column_ranges()
        row_ranges = self._get_row_ranges()
        header_data: list = []
        row = row_ranges['start']
        while row <= row_ranges['end']:
            header_row: list = []
            column = column_ranges['start']
            while column <= column_ranges['end']:
                merged_cell = None
                for cell in merged_cells:
                    cell_start, cell_end = cell.split(":")
                    if str(cell).startswith(f"{column}{row}:") or cell_start[0] <= column <= cell_end and int(cell_start[1]) <= row <= int(cell_end[1]):
                        merged_cell = cell
                        break
                if worksheet[f"{column}{row}"].value is not None:
                    if merged_cell is not None:
                        merged_cell = [item for item in merged_cell.split(":")]
                        header_row.append({
                            'name': worksheet[f"{column}{row}"].value,
                            'colspan': ord(merged_cell[1][0]) - ord(merged_cell[0][0]) + 1,
                            'rowspan': int(merged_cell[1][1:]) - int(merged_cell[0][1:]) + 1
                        })
                    else:
                        header_row.append({
                            'name': worksheet[f"{column}{row}"].value,
                            'colspan': 1,
                            'rowspan': 1
                        })
                column = chr(ord(column) + 1)
            header_data.append(header_row)
            row += 1
        workbook.close()
        return header_data

    def _get_merged_cells(self) -> list:
        workbook: Workbook = load_workbook(f"utils/excel/{self._type}.xlsx")
        worksheet: Worksheet = workbook["Жами"]
        merged_cells: MultiCellRange = worksheet.merged_cells
        merged_cells: Sequence = merged_cells.ranges
        merged_cells: list = [str(item) for item in merged_cells]
        merged_cells.sort()
        workbook.close()
        return merged_cells

    def _get_row_ranges(self) -> dict:
        merged_cells: list = self._get_merged_cells()
        row_ranges: dict = {
            'start': None,
            'end': None
        }
        workbook: Workbook = load_workbook(f"utils/excel/{self._type}.xlsx")
        worksheet: Worksheet = workbook["Жами"]
        for cell in merged_cells:
            cell_start, cell_end = (item for item in cell.split(":"))
            if int(cell_start[1:]) > 3 and not worksheet[cell_start].value.startswith("Жами"):
                if row_ranges['start'] is None:
                    row_ranges['start'] = int(cell_start[1:])
                else:
                    row_ranges['start'] = min(row_ranges['start'], int(cell_start[1:]))
                if row_ranges['end'] is None:
                    row_ranges['end'] = int(cell_end[1:])
                else:
                    row_ranges['end'] = max(row_ranges['end'], int(cell_end[1:]))
        workbook.close()
        return row_ranges

    def _get_table_data(self) -> dict:
        with open(self._table_data, "r", encoding="utf-8") as file:
            return load(file)[self._type]

    def _get_title(self) -> str:
        sql: SQLite = SQLite(f"database/excel/{self._type}.db")
        region: str = "Қорақалпоғистон Республикаси ва вилоятлар" \
            if self._region == "all" \
            else sql.select(
                table="region",
                fields=["name"],
                where=f"id = {self._region}",
                first=True
            )[0]
        sql.close()
        return f"{region}{self._table_data['title']}"

    def _insert_data(self, file: str, date: date) -> None:
        workbook: Workbook = load_workbook(file)
        sheets: list[str] = workbook.sheetnames
        with open("utils/regions.json", "r", encoding="utf-8") as file:
            regions: dict = load(file)
        for sheet in sheets:
            if sheet == "Жами":
                continue
            print(sheet)
            if sheet in regions:
                worksheet: Worksheet = workbook[sheet]
                column_ranges: dict = self._get_column_ranges()
                row_ranges: dict = self._get_row_ranges()
                column = column_ranges['start']
                row = row_ranges['end'] + 1
                sql: SQLite = SQLite(f"database/excel/{self._type}.db")
                while worksheet[f"{column}{row}"].value not in [None, "Жами"]:
                    for i in range(ord(column_ranges['start']) - ord("A") + 2, ord(column_ranges['end']) - ord("A") + 1):
                        column_keys_reverse: dict = {v: k for k, v in self._table_data['column_keys'].items()}
                        if i in column_keys_reverse:
                            district: str = worksheet[f"{chr(ord(column_ranges['start']) + 1)}{row}"].value
                            value: Any = worksheet[f"{chr(i + ord('A'))}{row}"].value or 0
                            sql.insert(
                                table="data",
                                name=district,
                                region=regions[sheet],
                                district=district,
                                date=date,
                                key=column_keys_reverse[i],
                                value=value,
                            )
                    row += 1
                    print(row)
                sql.close()
        workbook.close()

    def generate(self) -> Union[str, dict]:
        if '_format' not in self.__dict__:
            raise TableAttributeError("Table format is required.")
        try:
            generate_method = getattr(self, f"_generate_{self._format}")
            headers: list = self._get_headers()
            data: list = self._get_data()
            return generate_method(headers, data)
        except AttributeError:
            raise TableValueError(f"\"{self._format}\" format is not supported.")

    def get_column_ranges(self) -> dict:
        return self._get_column_ranges()

    def get_header_data(self) -> list:
        return self._get_header_data()

    def get_merged_cells(self) -> list:
        return self._get_merged_cells()
    
    def get_row_ranges(self) -> dict:
        return self._get_row_ranges()

    def get_title(self) -> str:
        return self._get_title()

    def insert_data(self, file: str, date: date) -> None:
        self._insert_data(file, date)
